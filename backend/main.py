from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# BASIC CONFIGURATION
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

# Vercel's deployed filesystem is read-only.
# /tmp is the writable directory.
OUTPUT_DIR = Path("/tmp/outputs")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


app = FastAPI(
    title="Survey Spreadsheet Validator",
    version="1.0.0",
)


# ============================================================
# CORS
# ============================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://upcensus-detector-zph2.vercel.app",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# COLUMN ALIASES
# ============================================================

LAT_ALIASES = {
    "location [latitude]",
    "location latitude",
    "latitude",
    "lat",
}

LON_ALIASES = {
    "location [longitude]",
    "location longitude",
    "longitude",
    "long",
    "lng",
}

MOBILE_ALIASES = {
    "mobile number",
    "mobile no",
    "mobile no.",
    "mobile",
    "contact number",
    "contact no",
    "contact no.",
    "phone number",
    "phone no",
    "phone",
}

OUTLET_ALIASES = {
    "outlet name",
    "shop name",
    "shop",
    "store name",
    "outlet",
}

IMAGE_HINTS = (
    "photo",
    "image",
    "picture",
    "pic",
    "screenshot",
)

# These metadata columns are NEVER cleaned, corrected, or Hindi-checked.
PROTECTED_METADATA_ALIASES = {
    "response id",
    "response start time",
    "response completion time",
    "ip address",
    "collector name",
}

def is_protected_metadata_column(column: Any) -> bool:
    return normalized_column_name(column) in PROTECTED_METADATA_ALIASES


# ============================================================
# TEXT HELPERS
# ============================================================

HINDI_RE = re.compile(r"[\u0900-\u097F]")


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass

    return str(value).strip()


def contains_hindi(value: Any) -> bool:
    return bool(HINDI_RE.search(clean_text(value)))


def normalized_column_name(value: Any) -> str:
    text = clean_text(value)

    text = (
        text.replace("\ufeff", " ")
        .replace("\u200b", " ")
        .replace("\u200c", " ")
        .replace("\u200d", " ")
        .replace("\xa0", " ")
        .replace("\r", " ")
        .replace("\n", " ")
        .replace("\t", " ")
    )

    text = text.lower().strip()

    text = re.sub(
        r"[^a-z0-9\u0900-\u097f]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def title_case_value(value: Any) -> Any:
    """
    Sentence-style capitalization for ordinary text:
    only the first letter of the first word is uppercase.
    """

    text = clean_text(value)

    if not text or contains_hindi(text):
        return value

    return text[:1].upper() + text[1:].lower()


def is_person_name_column(column: Any) -> bool:
    """
    Detect person/name columns.

    Every word in a person name gets its first letter capitalized.
    Shop/outlet/store names are deliberately excluded because those
    use sentence-style capitalization instead.
    """

    normalized = normalized_column_name(column)

    if "name" not in normalized:
        return False

    excluded = (
        "outlet",
        "shop",
        "store",
        "business",
    )

    return not any(
        word in normalized
        for word in excluded
    )


def person_name_value(value: Any) -> Any:
    """
    Person-name capitalization:
        rahul kumar -> Rahul Kumar
        RAHUL KUMAR -> Rahul Kumar

    Only the first character of each name word is uppercase.
    """

    text = clean_text(value)

    if not text or contains_hindi(text):
        return value

    return " ".join(
        word[:1].upper() + word[1:].lower()
        if word
        else word
        for word in text.split()
    )


def corrected_text_value(
    value: Any,
    column: Any,
) -> Any:
    """
    Apply the requested capitalization rule:
    - person/name columns: first letter of EVERY name word uppercase
    - outlet/shop and all other ordinary text: only FIRST word's
      first letter uppercase
    """

    if is_person_name_column(column):
        return person_name_value(value)

    return title_case_value(value)


# ============================================================
# COLUMN FINDER
# ============================================================

def find_column(
    df: pd.DataFrame,
    aliases: set[str],
) -> str | None:

    normalized_aliases = {
        normalized_column_name(alias)
        for alias in aliases
    }

    # Exact normalized match.
    for column in df.columns:
        normalized = normalized_column_name(column)

        if normalized in normalized_aliases:
            return column

    # Token fallback.
    for column in df.columns:
        normalized = normalized_column_name(column)
        words = set(normalized.split())

        if "mobile" in words and "number" in words:
            return column

        if "mobile" in words and "no" in words:
            return column

        if "contact" in words and "number" in words:
            return column

        if "phone" in words and "number" in words:
            return column

    # Substring fallback.
    for column in df.columns:
        normalized = normalized_column_name(column)

        if (
            "mobile" in normalized
            and (
                "number" in normalized
                or "no" in normalized
            )
        ):
            return column

        if (
            "contact" in normalized
            and "number" in normalized
        ):
            return column

        if (
            "phone" in normalized
            and "number" in normalized
        ):
            return column

    # Generic alias fallback.
    for column in df.columns:
        normalized = normalized_column_name(column)

        if any(
            alias in normalized
            for alias in normalized_aliases
        ):
            return column

    return None


# ============================================================
# IMAGE COLUMN DETECTION
# ============================================================

def find_image_columns(
    df: pd.DataFrame,
) -> list[str]:

    columns: list[str] = []

    for column in df.columns:
        name = normalized_column_name(column)

        if any(
            hint in name
            for hint in IMAGE_HINTS
        ):
            columns.append(column)

    return columns


def is_shop_image_column(
    column: Any,
) -> bool:

    normalized = normalized_column_name(column)

    return (
        "shop" in normalized
        or "outlet" in normalized
    )


# ============================================================
# INPUT FILE READER
# ============================================================

def read_input(
    path: Path,
) -> pd.DataFrame:

    suffix = path.suffix.lower()

    # --------------------------------------------------------
    # Read without assuming first row is header.
    # --------------------------------------------------------

    if suffix in {".xlsx", ".xls"}:
        raw_df = pd.read_excel(
            path,
            dtype=str,
            header=None,
        ).fillna("")

    elif suffix in {".csv", ".tsv"}:
        separator = (
            "\t"
            if suffix == ".tsv"
            else ","
        )

        raw_df = pd.read_csv(
            path,
            sep=separator,
            dtype=str,
            header=None,
            keep_default_na=False,
        ).fillna("")

    else:
        raise ValueError(
            "Unsupported file type."
        )

    if raw_df.empty:
        raise ValueError(
            "The uploaded spreadsheet is empty."
        )

    # --------------------------------------------------------
    # Header normalization.
    # --------------------------------------------------------

    def header_search_text(
        value: Any,
    ) -> str:

        text = clean_text(value)

        text = (
            text.replace("\ufeff", " ")
            .replace("\u200b", " ")
            .replace("\u200c", " ")
            .replace("\u200d", " ")
            .replace("\xa0", " ")
            .replace("\r", " ")
            .replace("\n", " ")
            .replace("\t", " ")
        )

        text = text.lower()

        text = re.sub(
            r"[^a-z0-9]+",
            " ",
            text,
        )

        return re.sub(
            r"\s+",
            " ",
            text,
        ).strip()

    # --------------------------------------------------------
    # Detect the real header row.
    # --------------------------------------------------------

    best_header_row: int | None = None
    best_score = -1

    rows_to_check = min(
        50,
        len(raw_df),
    )

    for row_index in range(rows_to_check):

        values = [
            header_search_text(value)
            for value in raw_df.iloc[
                row_index
            ].tolist()
        ]

        non_empty = [
            value
            for value in values
            if value
        ]

        if not non_empty:
            continue

        score = 0

        if any(
            value == "response id"
            or "response id" in value
            for value in non_empty
        ):
            score += 5

        if any(
            "mobile number" in value
            or (
                "mobile" in value
                and "number" in value
            )
            for value in non_empty
        ):
            score += 10

        if any(
            "outlet name" in value
            or "shop name" in value
            for value in non_empty
        ):
            score += 8

        if any(
            "latitude" in value
            for value in non_empty
        ):
            score += 5

        if any(
            "longitude" in value
            for value in non_empty
        ):
            score += 5

        if any(
            "collector name" in value
            for value in non_empty
        ):
            score += 3

        if any(
            "pin code" in value
            or "pincode" in value
            for value in non_empty
        ):
            score += 3

        if len(non_empty) >= 8:
            score += 2

        if score > best_score:
            best_score = score
            best_header_row = row_index

    if (
        best_header_row is None
        or best_score < 10
    ):
        raise ValueError(
            "Could not detect the survey header row. "
            "The spreadsheet must contain a 'Mobile Number' "
            "column/header."
        )

    # --------------------------------------------------------
    # Build dataframe using the detected header.
    # --------------------------------------------------------

    header_values = [
        clean_text(value)
        for value in raw_df.iloc[
            best_header_row
        ].tolist()
    ]

    # Make duplicate/blank headers safe.
    used_headers: dict[str, int] = {}
    final_headers: list[str] = []

    for index, value in enumerate(
        header_values
    ):

        header = value or f"Column {index + 1}"

        count = used_headers.get(
            header,
            0,
        )

        if count:
            final_header = (
                f"{header}_{count + 1}"
            )
        else:
            final_header = header

        used_headers[header] = count + 1
        final_headers.append(final_header)

    df = raw_df.iloc[
        best_header_row + 1:
    ].copy()

    df.columns = final_headers

    df = df.reset_index(
        drop=True
    )

    # Remove completely empty rows.
    non_empty_mask = df.apply(
        lambda row: any(
            clean_text(value)
            for value in row.tolist()
        ),
        axis=1,
    )

    df = df.loc[
        non_empty_mask
    ].reset_index(
        drop=True
    )

    return df


# ============================================================
# MOBILE NUMBER
# ============================================================

def mobile_digits(
    value: Any,
) -> str:

    text = clean_text(value)

    digits = re.sub(
        r"\D",
        "",
        text,
    )

    if (
        len(digits) == 12
        and digits.startswith("91")
    ):
        digits = digits[2:]

    return digits


def valid_indian_mobile(
    value: Any,
) -> bool:

    digits = mobile_digits(
        value
    )

    return bool(
        re.fullmatch(
            r"[6-9]\d{9}",
            digits,
        )
    )


def suspicious_mobile(
    value: Any,
) -> bool:

    digits = mobile_digits(
        value
    )

    if len(digits) != 10:
        return False

    if len(set(digits)) == 1:
        return True

    fake_patterns = {
        "0123456789",
        "1234567890",
        "9876543210",
        "0987654321",
    }

    return digits in fake_patterns


# ============================================================
# SHOP NAME NORMALIZATION
# ============================================================

def normalized_shop(
    value: Any,
) -> str:

    text = clean_text(
        value
    ).lower()

    text = re.sub(
        r"[^a-z0-9\u0900-\u097f]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


# ============================================================
# IMAGE REFERENCE
# ============================================================

def image_reference(
    value: Any,
) -> str:

    text = clean_text(
        value
    )

    if not text:
        return ""

    text = text.lower()

    text = re.sub(
        r"\s+",
        " ",
        text,
    ).strip()

    return text


def image_reason_label(
    value: Any,
) -> str:

    text = clean_text(
        value
    )

    if not text:
        return ""

    # Remove the survey instruction beginning with
    # ": Note : Interviewer ..."
    text = re.split(
        r"\s*:\s*note\b.*$",
        text,
        maxsplit=1,
        flags=re.IGNORECASE | re.DOTALL,
    )[0].strip()

    return text.rstrip(" :")


def looks_like_screenshot(
    value: Any,
) -> bool:

    text = clean_text(
        value
    ).lower()

    if not text:
        return False

    screenshot_terms = (
        "screenshot",
        "screen shot",
        "screen_shot",
        "screen-shot",
        "snip_",
        "snipping",
        "snip",
        "capture_",
        "capture-",
        "screen-record",
        "screen_record",
        "screen record",
        "screenrecord",
    )

    return any(
        term in text
        for term in screenshot_terms
    )


# ============================================================
# NUMERIC VALUE
# ============================================================

def numeric_value(
    value: Any,
) -> float | None:

    text = clean_text(
        value
    )

    if not text:
        return None

    try:
        return float(
            text.replace(",", "")
        )
    except ValueError:
        return None


# ============================================================
# EXCEL FORMATTING
# ============================================================

def format_worksheet(
    worksheet: Any,
) -> None:

    worksheet.freeze_panes = "A2"

    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1D4ED8",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 42

    for row in worksheet.iter_rows(
        min_row=2,
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for column_cells in worksheet.columns:

        column_letter = get_column_letter(
            column_cells[0].column
        )

        max_length = 0

        for cell in column_cells[:300]:

            value = (
                ""
                if cell.value is None
                else str(cell.value)
            )

            max_length = max(
                max_length,
                len(value),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = max(
            12,
            min(
                max_length + 2,
                45,
            ),
        )


# ============================================================
# MAIN VALIDATION
# ============================================================

def create_output(
    df: pd.DataFrame,
    output_path: Path,
) -> dict[str, Any]:

    mobile_col = find_column(
        df,
        MOBILE_ALIASES,
    )

    outlet_col = find_column(
        df,
        OUTLET_ALIASES,
    )

    lat_col = find_column(
        df,
        LAT_ALIASES,
    )

    lon_col = find_column(
        df,
        LON_ALIASES,
    )

    image_cols = find_image_columns(
        df
    )

    print(
        "\n========================================"
    )
    print(
        "IMPORTANT COLUMN DETECTION"
    )
    print(
        "========================================"
    )
    print(
        "Mobile Number:",
        repr(mobile_col),
    )
    print(
        "Outlet Name:",
        repr(outlet_col),
    )
    print(
        "Latitude:",
        repr(lat_col),
    )
    print(
        "Longitude:",
        repr(lon_col),
    )
    print(
        "Image Columns:",
        [
            repr(column)
            for column in image_cols
        ],
    )
    print(
        "========================================\n"
    )

    if mobile_col is None:

        detected_columns = [
            str(column)
            for column in df.columns
        ]

        raise ValueError(
            "Could not find the 'Mobile Number' column. "
            "Detected columns: "
            + ", ".join(
                detected_columns[:80]
            )
        )

    if outlet_col is None:
        raise ValueError(
            "Could not find the 'Outlet Name' column. "
            "Please check the spreadsheet header."
        )

    if lat_col is None:
        raise ValueError(
            "Could not find the Latitude column."
        )

    if lon_col is None:
        raise ValueError(
            "Could not find the Longitude column."
        )

    # ========================================================
    # DUPLICATE MOBILE NUMBERS
    # ========================================================

    mobile_series = df[
        mobile_col
    ].map(
        mobile_digits
    )

    mobile_counts = (
        mobile_series[
            mobile_series != ""
        ]
        .value_counts()
    )

    duplicate_mobiles = set(
        mobile_counts[
            mobile_counts > 1
        ].index
    )

    # --------------------------------------------------------
    # Duplicate mobile representative
    #
    # Normally the FIRST occurrence is kept.
    # If the first occurrence is missing latitude/longitude
    # and a later duplicate has both coordinates, keep that
    # later row instead. Every other duplicate is discarded.
    # --------------------------------------------------------

    mobile_occurrences: dict[str, list[int]] = {}

    for row_index, mobile in mobile_series.items():
        if mobile:
            mobile_occurrences.setdefault(
                mobile,
                [],
            ).append(row_index)

    duplicate_mobile_keep_row: dict[str, int] = {}

    for mobile in duplicate_mobiles:
        occurrences = mobile_occurrences.get(
            mobile,
            [],
        )

        if not occurrences:
            continue

        keep_row = occurrences[0]

        first_lat = numeric_value(
            df.iloc[keep_row][lat_col]
        )
        first_lon = numeric_value(
            df.iloc[keep_row][lon_col]
        )

        # Special case requested:
        # if first duplicate has missing coordinates, prefer
        # the first later duplicate that has both coordinates.
        if (
            first_lat is None
            or first_lon is None
        ):
            for candidate_row in occurrences[1:]:
                candidate_lat = numeric_value(
                    df.iloc[candidate_row][lat_col]
                )
                candidate_lon = numeric_value(
                    df.iloc[candidate_row][lon_col]
                )

                if (
                    candidate_lat is not None
                    and candidate_lon is not None
                ):
                    keep_row = candidate_row
                    break

        duplicate_mobile_keep_row[
            mobile
        ] = keep_row

    # ========================================================
    # IMAGE REFERENCE DUPLICATES
    # ========================================================

    image_occurrences: dict[
        str,
        list[tuple[int, str]],
    ] = {}

    for image_column in image_cols:

        for row_index, value in df[
            image_column
        ].items():

            reference = image_reference(
                value
            )

            if reference:
                image_occurrences.setdefault(
                    reference,
                    [],
                ).append(
                    (
                        row_index,
                        image_column,
                    )
                )

    duplicate_images = {
        reference
        for reference, occurrences
        in image_occurrences.items()
        if len(occurrences) > 1
    }

    # Same image in the SAME image column:
    # keep only the first occurrence.
    same_column_image_keep: dict[
        str,
        tuple[int, str],
    ] = {}

    # Same image used in DIFFERENT image columns:
    # discard every occurrence because the image was used
    # for the wrong/different product field as well.
    cross_column_duplicate_images: set[str] = set()

    for reference in duplicate_images:

        occurrences = image_occurrences[
            reference
        ]

        columns_used = {
            column
            for _, column in occurrences
        }

        if len(columns_used) > 1:
            cross_column_duplicate_images.add(
                reference
            )
        else:
            same_column_image_keep[
                reference
            ] = occurrences[0]

    # ========================================================
    # RESULT CONTAINERS
    # ========================================================

    correct_rows: list[
        dict[str, Any]
    ] = []

    discarded_rows: list[
        dict[str, Any]
    ] = []

    correct_row_indexes: list[int] = []
    discarded_row_indexes: list[int] = []

    # Store formatting information by original dataframe index.
    changed_columns_by_row: dict[
        int,
        set[str],
    ] = {}

    red_columns_by_row: dict[
        int,
        set[str],
    ] = {}

    # ========================================================
    # COUNTERS
    # ========================================================

    counters = {
        "Mobile Duplicates": 0,
        "Invalid Mobile": 0,
        "Suspicious Mobile": 0,
        "Blank Mobile Number": 0,
        "Duplicate Images": 0,
        "Screenshots": 0,
        "Missing Location": 0,
        "Invalid Coordinates": 0,
        "No Shop Image": 0,
        "Hindi Information": 0,
        "Corrected Cells": 0,
    }

    # ========================================================
    # PROCESS EVERY ROW
    # ========================================================

    for row_index, row in df.iterrows():

        reasons: list[str] = []

        changed_columns: set[str] = set()
        red_columns: set[str] = set()

        # ----------------------------------------------------
        # Hindi information
        # ----------------------------------------------------

        hindi_found = False

        for column in df.columns:

            if is_protected_metadata_column(column):
                continue

            if contains_hindi(
                row[column]
            ):

                hindi_found = True
                break

        if hindi_found:

            reasons.append(
                "Hindi Information Provided"
            )

            counters[
                "Hindi Information"
            ] += 1

        # ----------------------------------------------------
        # Mobile
        # ----------------------------------------------------

        mobile_value = clean_text(
            row[mobile_col]
        )

        mobile = mobile_digits(
            row[mobile_col]
        )

        if not mobile_value:

            reasons.append(
                "Blank Mobile Number"
            )

            counters[
                "Blank Mobile Number"
            ] += 1

            red_columns.add(
                str(mobile_col)
            )

        elif not valid_indian_mobile(
            row[mobile_col]
        ):

            reasons.append(
                "Invalid Indian Mobile Number"
            )

            counters[
                "Invalid Mobile"
            ] += 1

        if suspicious_mobile(
            row[mobile_col]
        ):

            reasons.append(
                "Suspicious Mobile Number"
            )

            counters[
                "Suspicious Mobile"
            ] += 1

        if (
            mobile
            and mobile in duplicate_mobiles
            and duplicate_mobile_keep_row.get(
                mobile
            ) != row_index
        ):

            reasons.append(
                "Duplicate Mobile Number"
            )

            counters[
                "Mobile Duplicates"
            ] += 1

        # ----------------------------------------------------
        # Latitude / Longitude
        #
        # IMPORTANT:
        # Missing coordinates are NOT discard conditions.
        # The row remains in Correct Data.
        # Only the blank cells are marked red.
        # ----------------------------------------------------

        lat = numeric_value(
            row[lat_col]
        )

        lon = numeric_value(
            row[lon_col]
        )

        if lat is None:

            counters[
                "Missing Location"
            ] += 1

            red_columns.add(
                str(lat_col)
            )

        if lon is None:

            counters[
                "Missing Location"
            ] += 1

            red_columns.add(
                str(lon_col)
            )

        if (
            lat is not None
            and lon is not None
        ):

            invalid_latitude = not (
                -90 <= lat <= 90
            )

            invalid_longitude = not (
                -180 <= lon <= 180
            )

            if (
                invalid_latitude
                or invalid_longitude
            ):

                reasons.append(
                    "Invalid Latitude/Longitude"
                )

                counters[
                    "Invalid Coordinates"
                ] += 1

        # ----------------------------------------------------
        # Images
        # ----------------------------------------------------

        for image_col in image_cols:

            reference = image_reference(
                row[image_col]
            )

            image_label = image_reason_label(
                image_col
            )

            is_shop_image = (
                is_shop_image_column(
                    image_col
                )
            )

            # Missing image is a discard condition ONLY
            # for shop/outlet image columns.
            if (
                is_shop_image
                and not reference
            ):

                reasons.append(
                    "No Shop Image"
                )

                counters[
                    "No Shop Image"
                ] += 1

                red_columns.add(
                    str(image_col)
                )

            if (
                reference
                and reference in duplicate_images
            ):
                occurrences = image_occurrences.get(
                    reference,
                    [],
                )

                current_occurrence = (
                    row_index,
                    image_col,
                )

                discard_duplicate_image = (
                    reference
                    in cross_column_duplicate_images
                    or (
                        same_column_image_keep.get(
                            reference
                        )
                        != current_occurrence
                    )
                )

                if discard_duplicate_image:
                    reasons.append(
                        "Duplicate Image Reference: "
                        + image_label
                    )

                    counters[
                        "Duplicate Images"
                    ] += 1

            if (
                reference
                and looks_like_screenshot(
                    reference
                )
            ):

                reasons.append(
                    "Screenshot Image: "
                    + image_label
                )

                counters[
                    "Screenshots"
                ] += 1

        # ----------------------------------------------------
        # Prepare row data.
        #
        # Normal Latin text is corrected using the requested capitalization rules.
        # Hindi is preserved for the discard report.
        # ----------------------------------------------------

        base: dict[str, Any] = {}

        for column in df.columns:

            column_name = str(
                column
            )

            original_value = row[
                column
            ]

            original_text = clean_text(
                original_value
            )

            # Never alter mobile numbers,
            # coordinates or image references.
            special_column = (
                column == mobile_col
                or column == lat_col
                or column == lon_col
                or column in image_cols
            )

            if (
                original_text
                and not special_column
                and not is_protected_metadata_column(column)
                and not contains_hindi(
                    original_text
                )
            ):

                corrected_value = (
                    corrected_text_value(
                        original_text,
                        column,
                    )
                )

                base[
                    column_name
                ] = corrected_value

                if (
                    corrected_value
                    != original_text
                ):

                    changed_columns.add(
                        column_name
                    )

            else:

                base[
                    column_name
                ] = original_value

        if changed_columns:
            counters[
                "Corrected Cells"
            ] += len(
                changed_columns
            )

        changed_columns_by_row[
            row_index
        ] = changed_columns

        red_columns_by_row[
            row_index
        ] = red_columns

        # ----------------------------------------------------
        # Latitude/longitude missing does NOT cause discard.
        # ----------------------------------------------------

        discard_reasons = list(
            dict.fromkeys(
                reasons
            )
        )

        if discard_reasons:

            base[
                "Discard Reason"
            ] = "; ".join(
                discard_reasons
            )

            discarded_rows.append(
                base
            )
            discarded_row_indexes.append(row_index)

        else:

            correct_rows.append(
                base
            )
            correct_row_indexes.append(row_index)

    # ========================================================
    # DATAFRAMES
    # ========================================================

    original_columns = [
        str(column)
        for column in df.columns
    ]

    correct_df = pd.DataFrame(
        correct_rows,
        columns=original_columns,
    )

    discard_df = pd.DataFrame(
        discarded_rows,
        columns=[
            *original_columns,
            "Discard Reason",
        ],
    )

    # ========================================================
    # CREATE EXCEL
    # ========================================================

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        correct_df.to_excel(
            writer,
            index=False,
            sheet_name="Correct Data",
        )

        discard_df.to_excel(
            writer,
            index=False,
            sheet_name="Discard Data",
        )

        workbook = writer.book

        correct_sheet = workbook[
            "Correct Data"
        ]

        discard_sheet = workbook[
            "Discard Data"
        ]

        format_worksheet(
            correct_sheet
        )

        format_worksheet(
            discard_sheet
        )

        # ----------------------------------------------------
        # Existing visual style for Correct Data.
        # Specific yellow/red cells are applied afterwards.
        # ----------------------------------------------------

        green_fill = PatternFill(
            fill_type="solid",
            fgColor="F0FDF4",
        )

        for row_number in range(
            2,
            correct_sheet.max_row + 1,
        ):

            for cell in correct_sheet[
                row_number
            ]:

                cell.fill = green_fill

        # ----------------------------------------------------
        # Discard reason styling.
        # ----------------------------------------------------

        if discard_sheet.max_column > 0:

            discard_reason_column = (
                discard_sheet.max_column
            )

            for row_number in range(
                2,
                discard_sheet.max_row + 1,
            ):

                cell = discard_sheet.cell(
                    row=row_number,
                    column=discard_reason_column,
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="FEE2E2",
                )

                cell.font = Font(
                    color="991B1B",
                    bold=True,
                )

        # ----------------------------------------------------
        # Yellow = validator corrected the cell.
        # Red = relevant blank cell.
        #
        # Because Correct Data can contain rows with only
        # missing latitude/longitude, red must be applied there.
        # ----------------------------------------------------

        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC",
        )

        red_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE",
        )

        red_font = Font(
            color="9C0006",
        )

        # Direct source-index -> output-row maps.
        # This avoids repeatedly scanning the Discard sheet.
        correct_output_row_by_source = {
            source_index: output_row
            for output_row, source_index in enumerate(
                correct_row_indexes,
                start=2,
            )
        }

        discard_output_row_by_source = {
            source_index: output_row
            for output_row, source_index in enumerate(
                discarded_row_indexes,
                start=2,
            )
        }

        for row_index in range(len(df)):

            if row_index in discard_output_row_by_source:
                target_sheet = discard_sheet
                target_row = discard_output_row_by_source[row_index]
            else:
                target_sheet = correct_sheet
                target_row = correct_output_row_by_source.get(row_index)

            if target_row is None:
                continue

            changed_columns = changed_columns_by_row.get(
                row_index, set()
            )
            red_columns = red_columns_by_row.get(
                row_index, set()
            )

            for column_index, column in enumerate(
                original_columns, start=1
            ):
                column_name = str(column)

                # Metadata columns remain untouched by validation highlighting.
                if is_protected_metadata_column(column):
                    continue

                cell = target_sheet.cell(
                    row=target_row,
                    column=column_index,
                )

                if column_name in changed_columns:
                    cell.fill = yellow_fill

                if column_name in red_columns:
                    cell.fill = red_fill
                    cell.font = red_font

    # ========================================================
    # RESPONSE
    # ========================================================

    return {
        "total_rows": len(df),
        "correct_rows": len(correct_df),
        "discarded_rows": len(discard_df),
        "output_filename": output_path.name,
        "checks": counters,
        "detected_columns": {
            "Mobile Number": str(
                mobile_col
            ),
            "Outlet Name": str(
                outlet_col
            ),
            "Latitude": str(
                lat_col
            ),
            "Longitude": str(
                lon_col
            ),
            "Image Columns": [
                str(column)
                for column in image_cols
            ],
        },
    }


# ============================================================
# HEALTH CHECK
# ============================================================

@app.get("/health")
def health() -> dict[str, str]:

    return {
        "status": "ok"
    }


# ============================================================
# VALIDATE FILE
# ============================================================

@app.post("/validate")
async def validate(
    file: UploadFile = File(...),
) -> dict[str, Any]:

    original_name = (
        file.filename
        or "survey_data.xlsx"
    )

    suffix = (
        Path(original_name)
        .suffix
        .lower()
    )

    allowed_extensions = {
        ".xlsx",
        ".xls",
        ".csv",
        ".tsv",
    }

    if suffix not in allowed_extensions:

        raise HTTPException(
            status_code=400,
            detail=(
                "Only XLSX, XLS, CSV and TSV "
                "files are supported."
            ),
        )

    job_id = uuid.uuid4().hex

    input_path = (
        OUTPUT_DIR
        / f"{job_id}{suffix}"
    )

    output_name = (
        f"Validated_"
        f"{Path(original_name).stem}_"
        f"{job_id[:8]}.xlsx"
    )

    output_path = (
        OUTPUT_DIR
        / output_name
    )

    try:

        file_content = await file.read()

        input_path.write_bytes(
            file_content
        )

        df = read_input(
            input_path
        )

        if df.empty:

            raise ValueError(
                "The uploaded spreadsheet "
                "contains no records."
            )

        result = create_output(
            df,
            output_path,
        )

        return result

    except HTTPException:
        raise

    except Exception as exc:

        if output_path.exists():
            output_path.unlink()

        raise HTTPException(
            status_code=400,
            detail=str(exc),
        ) from exc

    finally:

        if input_path.exists():
            input_path.unlink()


# ============================================================
# DOWNLOAD OUTPUT
# ============================================================

@app.get(
    "/download/{filename}"
)
def download(
    filename: str,
) -> FileResponse:

    safe_name = Path(
        filename
    ).name

    path = (
        OUTPUT_DIR
        / safe_name
    )

    if (
        not path.exists()
        or path.suffix.lower() != ".xlsx"
    ):

        raise HTTPException(
            status_code=404,
            detail="Output file not found.",
        )

    return FileResponse(
        path=path,
        filename=safe_name,
        media_type=(
            "application/vnd."
            "openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
